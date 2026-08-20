from __future__ import annotations

import json
import sqlite3
import zlib
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pytest

from manic.io.changelog_writer import generate_changelog
from manic.io.compound_reader import read_compound
from manic.io.compounds_import import detect_compound_list_format, import_compound_excel
from manic.io.data_exporter import DataExporter
from manic.io.data_provider import DataProvider
from manic.io.eic_importer import _iter_compounds
from manic.io.eic_reader import read_eic
from manic.models import database
from manic.models import session_export
from manic.models.analysis import AnalysisMode, IonRole
from manic.processors import chromatographic_peak_deconvolution as deconv
from manic.processors import integration as integration_module
from manic.processors.chromatographic_peak_deconvolution import (
    ChannelDeconvolution,
    ChannelDeconvolutionBundle,
    EICChromatographicPeakDeconvolutionResult,
    deconvolve_eic,
)
from manic.processors.integration import calculate_peak_areas
from manic.validation.unlabelled_identity import IdentityStatus


SCHEMA = Path(__file__).parent.parent / "src" / "manic" / "models" / "schema.sql"


@pytest.fixture
def unlabelled_db(tmp_path, monkeypatch):
    db_path = tmp_path / "unlabelled.db"
    monkeypatch.setattr(database, "DB_FILE", db_path)
    with sqlite3.connect(db_path) as conn:
        conn.executescript(SCHEMA.read_text(encoding="utf-8"))
    return db_path


def _import_targets(tmp_path, **row) -> int:
    """Write a one-row unlabelled compound list and import it."""
    path = tmp_path / "targets.csv"
    pd.DataFrame([row]).to_csv(path, index=False)
    return import_compound_excel(path, AnalysisMode.UNLABELLED)


def _gaussian(time, center, width, height):
    return height * np.exp(-0.5 * ((time - center) / width) ** 2)


def _enable_deconvolution(compound_name: str, level: str = "4") -> None:
    with database.get_connection() as conn:
        conn.execute(
            "UPDATE compounds SET deconvolution_level = ? WHERE compound_name = ?",
            (level, compound_name),
        )


def _insert_eic(sample: str, compound: str, time, matrix, rt_window: float) -> None:
    """Insert a sample and its (compressed) multi-channel EIC trace."""
    with database.get_connection() as conn:
        conn.execute(
            "INSERT OR IGNORE INTO samples (sample_name, file_name) VALUES (?, ?)",
            (sample, f"/fake/{sample}.cdf"),
        )
        conn.execute(
            "INSERT INTO eic (sample_name, compound_name, x_axis, y_axis, rt_window) "
            "VALUES (?, ?, ?, ?, ?)",
            (
                sample,
                compound,
                zlib.compress(np.asarray(time, dtype=np.float64).tobytes()),
                zlib.compress(np.asarray(matrix, dtype=np.float64).ravel().tobytes()),
                rt_window,
            ),
        )


def test_detect_format_recognises_gv3_csv(tmp_path):
    path = tmp_path / "gv3.csv"
    path.write_text(
        "name,tR,lOffset,rOffset,QIon,ValIon1,ValIon2,tR_Window\n"
        "A,1.0,0.1,0.1,100,150,200,0.1\n"
    )
    assert detect_compound_list_format(path) is AnalysisMode.UNLABELLED


def test_detect_format_recognises_gv5_xlsx(tmp_path):
    path = tmp_path / "gv5.xlsx"
    pd.DataFrame(
        {
            "name": ["Pyruvate"],
            "tR": [6.37],
            "Mass0": [174],
            "LabelAtoms": [3],
        }
    ).to_excel(path, index=False)
    assert detect_compound_list_format(path) is AnalysisMode.LABELLED


def test_detect_format_normalises_header_variants(tmp_path):
    path = tmp_path / "variants.csv"
    path.write_text("Name,tR,Quant Ion,Val Ion 1\nA,1.0,100,150\n")
    assert detect_compound_list_format(path) is AnalysisMode.UNLABELLED


def test_detect_format_returns_none_for_unrecognised_headers(tmp_path):
    path = tmp_path / "other.csv"
    path.write_text("a,b,c\n1,2,3\n")
    assert detect_compound_list_format(path) is None


def test_detect_format_returns_none_for_missing_file(tmp_path):
    assert detect_compound_list_format(tmp_path / "missing.csv") is None


def test_existing_database_migrates_targeted_schema(tmp_path, monkeypatch):
    db_path = tmp_path / "legacy.db"
    monkeypatch.setattr(database, "DB_FILE", db_path)

    schema_text = SCHEMA.read_text(encoding="utf-8")
    schema_text = schema_text.replace(
        "    rt_tolerance  REAL,  -- Optional identity-QC tolerance for targeted analysis (minutes)\n",
        "",
    )
    ions_start = schema_text.index("-- Explicit diagnostic ions")
    samples_start = schema_text.index("-- Samples", ions_start)
    legacy_schema = schema_text[:ions_start] + schema_text[samples_start:]
    with sqlite3.connect(db_path) as conn:
        conn.executescript(legacy_schema)
        conn.execute(
            "INSERT INTO compounds (compound_name, retention_time, mass0) "
            "VALUES ('Existing', 1.2, 100)"
        )

    database.init_db()

    with sqlite3.connect(db_path) as conn:
        columns = {
            row[1] for row in conn.execute("PRAGMA table_info(compounds)")
        }
        tables = {
            row[0]
            for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            )
        }
        existing = conn.execute(
            "SELECT compound_name FROM compounds WHERE compound_name = 'Existing'"
        ).fetchone()

    assert "rt_tolerance" in columns
    assert "compound_ions" in tables
    assert existing == ("Existing",)


def test_unlabelled_compound_import_stores_arbitrary_ion_channels(
    unlabelled_db, tmp_path
):
    count = _import_targets(
        tmp_path,
        name="Citrate 4TMS",
        tR=12.4,
        lOffset=0.15,
        rOffset=0.2,
        QIon=273,
        ValIon1=147,
        ValIon2=73,
        **{"Qualifier 1 Ratio": 0.42, "Qualifier 1 Tolerance": 0.25},
    )

    assert count == 1
    compound = read_compound("Citrate 4TMS")
    assert compound.is_unlabelled_target
    assert compound.label_atoms == 0
    assert compound.mass0 == 273
    assert compound.channel_count == 3
    assert [channel.mz for channel in compound.analysis_channels] == [273, 147, 73]
    assert [channel.role for channel in compound.analysis_channels] == [
        IonRole.QUANTIFIER,
        IonRole.QUALIFIER,
        IonRole.QUALIFIER,
    ]
    assert compound.deconvolution_level == "off"
    assert compound.baseline_correction == 1
    assert compound.analysis_channels[1].expected_ratio == pytest.approx(0.42)
    assert compound.analysis_channels[1].ratio_tolerance == pytest.approx(0.25)


def test_unlabelled_import_requires_quantifier_and_qualifier_columns(
    unlabelled_db, tmp_path
):
    compound_list = tmp_path / "labelled-shaped.csv"
    pd.DataFrame(
        [
            {
                "name": "Alanine",
                "tR": 3.2,
                "mass0": 116,
                "labelAtoms": 3,
                "lOffset": 0.1,
                "rOffset": 0.1,
            }
        ]
    ).to_csv(compound_list, index=False)

    with pytest.raises(ValueError, match="missing unlabelled column"):
        import_compound_excel(
            compound_list,
            analysis_mode=AnalysisMode.UNLABELLED,
        )


def test_data_provider_integrates_all_channels_but_quantifies_quantifier(
    unlabelled_db, tmp_path
):
    _import_targets(
        tmp_path,
        name="Target",
        tR=1.0,
        lOffset=1.1,
        rOffset=1.1,
        quant_ion=217,
        qualifier_ion_1=147,
        qualifier_ion_2=73,
    )

    with database.get_connection() as conn:
        extraction_target = list(_iter_compounds(conn))[0]
    assert extraction_target.required_rt_window > 1.1
    # areas[0] is only the quantifier if channel 0 is the quantifier in both
    # the extraction target and the channel model. Pin that invariant here:
    # quantification and identity QC silently corrupt if it ever breaks.
    assert extraction_target.target_mzs[0] == pytest.approx(217.0)
    channels = read_compound("Target").analysis_channels
    assert channels[0].role is IonRole.QUANTIFIER
    assert [c.mz for c in channels] == pytest.approx([217.0, 147.0, 73.0])

    time = [0.0, 1.0, 2.0]
    matrix = [
        [0.0, 10.0, 0.0],
        [0.0, 4.0, 0.0],
        [0.0, 2.0, 0.0],
    ]
    _insert_eic("S1", "Target", time, matrix, rt_window=1.0)

    provider = DataProvider()
    round_trip = read_eic(
        "S1",
        read_compound("Target"),
        use_corrected=False,
    )
    assert round_trip.intensity.shape == (3, 3)
    assert np.array_equal(round_trip.intensity, matrix)

    assert provider.get_compound_areas("S1", "Target") == pytest.approx(
        [10.0, 4.0, 2.0]
    )
    assert provider.get_compound_total_area("S1", "Target") == pytest.approx(10.0)
    qc = provider.assess_unlabelled_identity("S1", "Target")
    assert qc.status is IdentityStatus.NOT_ASSESSED
    assert qc.observed_rt == pytest.approx(1.0)
    assert [ratio.observed_ratio for ratio in qc.qualifier_ratios] == pytest.approx(
        [0.4, 0.2]
    )


def test_manual_integration_rt_is_used_for_identity_reference(
    unlabelled_db, tmp_path
):
    _import_targets(
        tmp_path,
        name="Target",
        tR=1.0,
        **{"tR Window": 0.05},
        lOffset=0.1,
        rOffset=0.1,
        quant_ion=217,
        qualifier_ion_1=147,
        **{"Qualifier 1 Ratio": 0.5, "Qualifier 1 Tolerance": 0.2},
    )

    _insert_eic(
        "S1",
        "Target",
        [1.0, 1.1, 1.2, 1.3],
        [
            [0.0, 2.0, 10.0, 0.0],
            [0.0, 1.0, 5.0, 0.0],
        ],
        rt_window=0.3,
    )
    with database.get_connection() as conn:
        conn.execute(
            """
            INSERT INTO session_activity
                (compound_name, sample_name, retention_time, loffset, roffset)
            VALUES (?, ?, ?, ?, ?)
            """,
            ("Target", "S1", 1.2, 0.11, 0.11),
        )

    qc = DataProvider().assess_unlabelled_identity("S1", "Target")

    assert qc.observed_rt == pytest.approx(1.2)
    assert qc.rt_error == pytest.approx(0.0)
    assert qc.rt_passed is True
    assert qc.status is IdentityStatus.SUPPORTED


def test_unlabelled_session_round_trip_preserves_mode_and_ions(
    unlabelled_db, tmp_path
):
    _import_targets(
        tmp_path,
        name="Target",
        tR=1.0,
        lOffset=0.1,
        rOffset=0.1,
        QIon=217,
        ValIon1=147,
        **{
            "tR Window": 0.08,
            "Qualifier 1 Ratio": 0.4,
            "Qualifier 1 Tolerance": 0.25,
            "Amount in StdMix": 2.5,
            "Int Std amount": 1.2,
            "MM Files": "*_MM_*",
        },
    )

    assert session_export.export_session_method(
        str(tmp_path / "method"),
        AnalysisMode.UNLABELLED,
    )
    method_path = tmp_path / "manic_session_export" / "method.json"
    exported = json.loads(method_path.read_text(encoding="utf-8"))
    assert exported["analysis_mode"] == "unlabelled"
    assert exported["compounds"][0]["ions"][1]["expected_ratio"] == pytest.approx(0.4)
    assert exported["compounds"][0]["amount_in_std_mix"] == pytest.approx(2.5)
    assert exported["compounds"][0]["int_std_amount"] == pytest.approx(1.2)
    assert exported["compounds"][0]["mm_files"] == "*_MM_*"

    with database.get_connection() as conn:
        conn.execute(
            "UPDATE compound_ions SET expected_ratio = 0.9 "
            "WHERE compound_name = 'Target' AND role = 'qualifier'"
        )
        conn.execute(
            "UPDATE compounds SET amount_in_std_mix = NULL, "
            "int_std_amount = NULL, mm_files = NULL "
            "WHERE compound_name = 'Target'"
        )

    ok, _ = session_export.import_session_overrides(
        str(method_path),
        expected_mode=AnalysisMode.UNLABELLED,
    )
    assert ok
    assert read_compound("Target").analysis_channels[1].expected_ratio == pytest.approx(
        0.4
    )
    with database.get_connection() as conn:
        metadata = conn.execute(
            "SELECT amount_in_std_mix, int_std_amount, mm_files "
            "FROM compounds WHERE compound_name = 'Target'"
        ).fetchone()
    assert metadata["amount_in_std_mix"] == pytest.approx(2.5)
    assert metadata["int_std_amount"] == pytest.approx(1.2)
    assert metadata["mm_files"] == "*_MM_*"

    valid, error = session_export.validate_method_file(
        str(method_path),
        expected_mode=AnalysisMode.LABELLED,
    )
    assert not valid
    assert "Start a new analysis" in error


def test_unlabelled_excel_export_uses_targeted_sheets(unlabelled_db, tmp_path):
    _import_targets(
        tmp_path,
        name="Target",
        tR=1.0,
        lOffset=1.1,
        rOffset=1.1,
        QIon=217,
        ValIon1=147,
    )

    _insert_eic(
        "S1",
        "Target",
        [0.0, 1.0, 2.0],
        [[0.0, 10.0, 0.0], [0.0, 4.0, 0.0]],
        rt_window=1.1,
    )

    export_path = tmp_path / "targeted.xlsx"
    assert DataExporter(AnalysisMode.UNLABELLED).export_to_excel(str(export_path))

    workbook = openpyxl.load_workbook(export_path, data_only=True)
    assert workbook.sheetnames == [
        "Targeted Results",
        "Qualifier QC",
        "Targeted Method",
    ]
    result = workbook["Targeted Results"]
    assert result["D2"].value == pytest.approx(10.0)
    assert result["H2"].value == IdentityStatus.NOT_ASSESSED.value

    with database.get_connection() as conn:
        conn.execute(
            """
            INSERT INTO session_activity
                (compound_name, sample_name, retention_time, loffset, roffset)
            VALUES (?, ?, ?, ?, ?)
            """,
            ("Target", "S1", 1.25, 1.1, 1.1),
        )
    export_path = tmp_path / "targeted_session_tr.xlsx"
    assert DataExporter(AnalysisMode.UNLABELLED).export_to_excel(str(export_path))
    method = openpyxl.load_workbook(export_path, data_only=True)["Targeted Method"]
    rows = list(method.iter_rows(min_row=1, max_row=40, values_only=True))
    assert any(row and row[0] == "Ion definitions" for row in rows)
    assert any(row and row[0] == "Current tR" for row in rows)
    ion_header = next(row for row in rows if row and row[0] == "Compound")
    assert ion_header[1] == "Role"
    assert "tR (min)" not in ion_header
    current_header = next(row for row in rows if row and row[0] == "Sample")
    assert current_header[:3] == ("Sample", "Compound", "tR (min)")
    current_row = next(
        row for row in rows if row and row[0] == "S1" and row[1] == "Target"
    )
    assert current_row[2] == pytest.approx(1.25)


def _unlabelled_mixed_bundle(time, *, failed_index: int):
    fitted = deconvolve_eic(
        time,
        _gaussian(time, 5.0, 0.08, 12.0),
        retention_time=5.0,
        loffset=0.4,
        roffset=0.4,
        stringency="4",
    )
    assert fitted.model is not None
    failed = EICChromatographicPeakDeconvolutionResult(
        selected=np.full(time.size, 3.0),
        selected_mask=np.asarray(fitted.selected_mask, dtype=bool),
        excluded=[],
        excluded_masks=[],
        selected_center=5.0,
        component_centers=[5.0],
        model=None,
    )
    channels = [
        ChannelDeconvolution(index=0, result=fitted),
        ChannelDeconvolution(index=1, result=fitted),
    ]
    channels[failed_index] = ChannelDeconvolution(index=failed_index, result=failed)
    return ChannelDeconvolutionBundle(time=time, channels=tuple(channels))


def test_deconvolution_on_quantifies_q_only_and_pairs_vq(unlabelled_db, tmp_path):
    deconv._fit_joint_component_model_cached.cache_clear()
    deconv._fit_single_component_model_cached.cache_clear()
    _import_targets(
        tmp_path,
        name="Target",
        tR=7.0,
        lOffset=4.0,
        rOffset=4.0,
        QIon=217,
        ValIon1=147,
        **{"Qualifier 1 Ratio": 0.4, "Qualifier 1 Tolerance": 0.25},
    )
    _enable_deconvolution("Target")

    time = np.linspace(0.0, 10.0, 201)
    matrix = [
        _gaussian(time, 4.0, 0.25, 10.0) + _gaussian(time, 7.0, 0.25, 6.0),
        _gaussian(time, 7.0, 0.25, 2.4),
    ]
    _insert_eic("S1", "Target", time, matrix, rt_window=4.0)

    provider = DataProvider()
    areas = provider.get_compound_areas("S1", "Target")
    modelled = calculate_peak_areas(
        time,
        np.asarray(matrix, dtype=np.float64).ravel(),
        0,
        7.0,
        4.0,
        4.0,
        channel_count=2,
        baseline_correction=True,
        chromatographic_peak_deconvolution_stringency="4",
    )
    raw_window = calculate_peak_areas(
        time,
        np.asarray(matrix, dtype=np.float64).ravel(),
        0,
        7.0,
        4.0,
        4.0,
        channel_count=2,
        baseline_correction=True,
        chromatographic_peak_deconvolution_stringency="off",
    )

    assert areas == pytest.approx(modelled)
    assert areas[0] < raw_window[0]
    assert provider.get_compound_total_area("S1", "Target") == pytest.approx(areas[0])
    assert provider.get_compound_total_area("S1", "Target") != pytest.approx(sum(areas))

    qc = provider.assess_unlabelled_identity("S1", "Target")
    assert qc.quantifier_area == pytest.approx(areas[0])
    assert [ratio.observed_ratio for ratio in qc.qualifier_ratios] == pytest.approx(
        [areas[1] / areas[0]]
    )
    assert qc.observed_rt == pytest.approx(4.0, abs=0.05)

    bulk = DataProvider().get_sample_raw_data("S1")["Target"]
    assert bulk == pytest.approx(areas)

    export_path = tmp_path / "deconv.xlsx"
    assert DataExporter(AnalysisMode.UNLABELLED).export_to_excel(str(export_path))
    result = openpyxl.load_workbook(export_path, data_only=True)["Targeted Results"]
    assert result["D2"].value == pytest.approx(areas[0])


def test_raw_calibrated_expected_ratio_fails_after_deconvolution_on(
    unlabelled_db, tmp_path
):
    deconv._fit_joint_component_model_cached.cache_clear()
    deconv._fit_single_component_model_cached.cache_clear()
    time = np.linspace(0.0, 10.0, 201)
    matrix = np.vstack(
        [
            _gaussian(time, 4.0, 0.25, 10.0) + _gaussian(time, 7.0, 0.25, 6.0),
            _gaussian(time, 7.0, 0.25, 2.4),
        ]
    )
    raw = calculate_peak_areas(
        time,
        matrix.ravel(),
        0,
        7.0,
        4.0,
        4.0,
        channel_count=2,
        baseline_correction=True,
        chromatographic_peak_deconvolution_stringency="off",
    )
    raw_ratio = raw[1] / raw[0]
    _import_targets(
        tmp_path,
        name="Target",
        tR=7.0,
        lOffset=4.0,
        rOffset=4.0,
        QIon=217,
        ValIon1=147,
        **{"Qualifier 1 Ratio": raw_ratio, "Qualifier 1 Tolerance": 0.25},
    )
    _insert_eic("S1", "Target", time, matrix, rt_window=4.0)

    qc_off = DataProvider().assess_unlabelled_identity("S1", "Target")
    assert qc_off.qualifier_ratios[0].passed is True
    assert qc_off.qualifier_ratios[0].observed_ratio == pytest.approx(raw_ratio)

    _enable_deconvolution("Target")
    qc_on = DataProvider().assess_unlabelled_identity("S1", "Target")
    assert read_compound("Target").baseline_correction == 1
    assert qc_on.qualifier_ratios[0].observed_ratio != pytest.approx(raw_ratio, rel=0.1)
    assert qc_on.qualifier_ratios[0].passed is False
    assert qc_on.status is IdentityStatus.REVIEW_REQUIRED


@pytest.mark.parametrize("failed_index", [0, 1])
def test_deconvolution_on_falls_back_when_any_ion_fails(
    unlabelled_db, tmp_path, monkeypatch, failed_index
):
    _import_targets(
        tmp_path,
        name="Target",
        tR=5.0,
        lOffset=0.4,
        rOffset=0.4,
        QIon=217,
        ValIon1=147,
        **{"Qualifier 1 Ratio": 0.25, "Qualifier 1 Tolerance": 0.25},
    )
    _enable_deconvolution("Target")

    time = np.linspace(4.0, 6.0, 81)
    matrix = np.vstack(
        [
            _gaussian(time, 5.0, 0.08, 12.0),
            np.full(time.size, 3.0),
        ]
    )
    _insert_eic("S1", "Target", time, matrix, rt_window=0.4)
    monkeypatch.setattr(
        integration_module,
        "deconvolve_channel_matrix",
        lambda *args, **kwargs: _unlabelled_mixed_bundle(
            time, failed_index=failed_index
        ),
    )

    expected = calculate_peak_areas(
        time,
        matrix.ravel(),
        0,
        5.0,
        0.4,
        0.4,
        channel_count=2,
        baseline_correction=True,
        chromatographic_peak_deconvolution_stringency="off",
    )
    provider = DataProvider()
    areas = provider.get_compound_areas("S1", "Target")
    assert areas == pytest.approx(expected)
    assert provider.get_compound_total_area("S1", "Target") == pytest.approx(expected[0])

    qc = provider.assess_unlabelled_identity("S1", "Target")
    assert qc.quantifier_area == pytest.approx(expected[0])
    assert [ratio.observed_ratio for ratio in qc.qualifier_ratios] == pytest.approx(
        [expected[1] / expected[0]]
    )
    assert DataProvider().get_sample_raw_data("S1")["Target"] == pytest.approx(expected)


def test_unlabelled_component_fallback_keeps_qv_channels(unlabelled_db):
    time = np.array([0.0, 1.0, 2.0])
    intensity = np.array([[0.0, 10.0, 0.0], [0.0, 4.0, 0.0]]).ravel()
    row = {
        "label_atoms": 0,
        "formula": "",
        "channel_count": 2,
        "retention_time": 1.0,
        "loffset": 1.1,
        "roffset": 1.1,
        "deconvolution_level": "off",
        "deconvolution_fit_type": "auto",
        "deconvolution_noise_gate": "balanced",
    }
    provider = DataProvider()
    raw, corrected = provider._calculate_raw_and_corrected_areas_from_raw_component(
        time,
        intensity,
        row,
        use_legacy=False,
        baseline_correction=False,
    )
    assert raw == pytest.approx([10.0, 4.0])
    assert corrected == pytest.approx(raw)
    assert corrected is not raw
    assert provider._calculate_corrected_areas_from_raw_component(
        time,
        intensity,
        row,
        use_legacy=False,
        baseline_correction=False,
    ) == pytest.approx([10.0, 4.0])


def test_unlabelled_changelog_distinguishes_chromatographic_deconvolution(
    unlabelled_db, tmp_path
):
    _import_targets(
        tmp_path,
        name="Target",
        tR=1.0,
        lOffset=0.1,
        rOffset=0.1,
        QIon=217,
        ValIon1=147,
    )
    export_path = tmp_path / "targeted.xlsx"
    generate_changelog(
        str(export_path),
        internal_standard=None,
        use_legacy_integration=False,
        analysis_mode=AnalysisMode.UNLABELLED,
    )
    changelog = next(tmp_path.glob("changelog_*.md")).read_text(encoding="utf-8")
    assert "Natural-isotope correction is not applied" in changelog
    assert "isotopologue deconvolution are not applied" not in changelog
    assert "Chromatographic peak deconvolution is off unless enabled per compound" in changelog
    assert "Deconvolution" in changelog
