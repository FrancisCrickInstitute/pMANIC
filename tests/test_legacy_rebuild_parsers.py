import tempfile
from pathlib import Path

import numpy as np
from openpyxl import Workbook

from manic.io.legacy_raw_values_reader import read_raw_values_workbook
from manic.io.in_memory_provider import InMemoryDataProvider


def write_raw_values_xlsx(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Raw Values'

    # Headers
    ws.cell(row=1, column=1, value='Compound Name')
    ws.cell(row=2, column=1, value='Mass')
    ws.cell(row=3, column=1, value='Isotope')
    ws.cell(row=4, column=1, value='tR')
    ws.cell(row=1, column=2, value=None)
    ws.cell(row=2, column=2, value=None)
    ws.cell(row=3, column=2, value=None)
    ws.cell(row=4, column=2, value=None)

    # Two compounds: A with 1 label (M+0,M+1), B with unlabeled only
    ws.cell(row=1, column=3, value='A')
    ws.cell(row=2, column=3, value=100.0)
    ws.cell(row=3, column=3, value=0)
    ws.cell(row=4, column=3, value=1.0)

    ws.cell(row=1, column=4, value='A')
    ws.cell(row=2, column=4, value=100.0)
    ws.cell(row=3, column=4, value=1)
    ws.cell(row=4, column=4, value=1.0)

    ws.cell(row=1, column=5, value='B')
    ws.cell(row=2, column=5, value=50.0)
    ws.cell(row=3, column=5, value=0)
    ws.cell(row=4, column=5, value=2.0)

    # Sample rows
    ws.cell(row=5, column=1, value=None)
    ws.cell(row=5, column=2, value='S1')
    ws.cell(row=5, column=3, value=10.0)  # A M+0
    ws.cell(row=5, column=4, value=2.0)   # A M+1
    ws.cell(row=5, column=5, value=5.0)   # B M+0

    ws.cell(row=6, column=1, value=None)
    ws.cell(row=6, column=2, value='S2')
    ws.cell(row=6, column=3, value=20.0)
    ws.cell(row=6, column=4, value=3.0)
    ws.cell(row=6, column=5, value=7.0)

    wb.save(path)


def test_read_raw_values_workbook_roundtrip(tmp_path: Path):
    xlsx_path = tmp_path / 'raw_values.xlsx'
    write_raw_values_xlsx(str(xlsx_path))
    samples, raw = read_raw_values_workbook(str(xlsx_path))
    assert samples == ['S1', 'S2']
    assert raw['S1']['A'] == [10.0, 2.0]
    assert raw['S1']['B'] == [5.0]
    assert raw['S2']['A'] == [20.0, 3.0]
    assert raw['S2']['B'] == [7.0]


def test_in_memory_provider_corrected_vector(tmp_path: Path):
    # Compounds: A has label_atoms=1 (2 iso), B is unlabeled
    compounds = [
        {
            'compound_name': 'A',
            'retention_time': 1.0,
            'mass0': 100.0,
            'loffset': 0.1,
            'roffset': 0.1,
            'label_atoms': 1,
            'formula': 'C1H4',
            'label_type': 'C',
            'tbdms': 0,
            'meox': 0,
            'me': 0,
        },
        {
            'compound_name': 'B',
            'retention_time': 2.0,
            'mass0': 50.0,
            'loffset': 0.1,
            'roffset': 0.1,
            'label_atoms': 0,
            'formula': None,
            'label_type': 'C',
            'tbdms': 0,
            'meox': 0,
            'me': 0,
        },
    ]
    samples = ['S1']
    raw = {'S1': {'A': [10.0, 2.0], 'B': [5.0]}}

    prov = InMemoryDataProvider(compounds, samples, raw)

    corr = prov.get_sample_corrected_data('S1')
    # A: two isotopologues, non-negative, some total (approximate mode)
    assert isinstance(corr['A'], list) and len(corr['A']) == 2
    assert all(v >= 0 for v in corr['A'])
    # B: unlabeled, copied as-is
    assert corr['B'] == [5.0]
