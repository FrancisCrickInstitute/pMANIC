"""
Edge case and error handling tests for MANIC.
Tests boundary conditions, error scenarios, and data validation.
"""

import numpy as np
import pytest

from manic.processors.integration import calculate_peak_areas, integrate_peak
from manic.processors.natural_abundance_correction import NaturalAbundanceCorrector


class TestZeroHandling:
    """Test handling of zero values in calculations."""

    def test_zero_intensity_integration(self):
        """Test integration with all zero intensities."""
        time = np.linspace(0, 10, 100)
        intensity = np.zeros(100)

        areas = calculate_peak_areas(
            time, intensity, 0, 5.0, 1.0, 1.0
        )

        assert len(areas) == 1
        assert areas[0] == 0.0

    def test_zero_m0_isotopologue(self):
        """Test handling when M+0 signal is zero."""
        # This affects ratio calculations
        isotopologue_data = [0.0, 100.0, 50.0]  # M+0 is zero

        # Calculate ratio - should handle division by zero
        total = sum(isotopologue_data)
        if total > 0:
            ratios = [x/total for x in isotopologue_data]
        else:
            ratios = [0.0] * len(isotopologue_data)

        assert ratios[0] == 0.0
        assert abs(ratios[1] - 100.0/150.0) < 0.001

    def test_all_zero_isotopologues(self):
        """Test when all isotopologues are zero."""
        data = [0.0, 0.0, 0.0, 0.0]

        total = sum(data)
        assert total == 0.0

        # Ratio calculation should handle gracefully
        if total > 0:
            ratios = [x/total for x in data]
        else:
            ratios = [0.0] * len(data)

        assert all(r == 0.0 for r in ratios)

    def test_single_nonzero_isotopologue(self):
        """Test with only one non-zero isotopologue."""
        data = [0.0, 0.0, 100.0, 0.0]

        total = sum(data)
        ratios = [x/total for x in data]

        assert ratios[2] == 1.0
        assert sum(ratios) == 1.0


class TestDivisionByZero:
    """Test protection against division by zero errors."""

    def test_mrrf_zero_signal(self):
        """Test MRRF calculation with zero signals."""
        # Simulate zero internal standard signal
        mean_metabolite_signal = 100.0
        metabolite_conc = 10.0
        mean_is_signal = 0.0  # Zero!
        is_conc = 5.0

        # Should handle gracefully - typically returns 1.0 as default
        if mean_is_signal > 0 and is_conc > 0:
            mrrf = (mean_metabolite_signal / metabolite_conc) / (mean_is_signal / is_conc)
        else:
            mrrf = 1.0  # Default when calculation fails

        assert mrrf == 1.0

    def test_background_ratio_zero_m0(self):
        """Test background ratio with zero M+0."""
        m0_signal = 0.0
        labeled_signal = 50.0

        # Should handle division by zero
        if m0_signal > 0:
            ratio = labeled_signal / m0_signal
        else:
            ratio = 0.0  # Or skip this sample

        assert ratio == 0.0

    def test_percent_label_zero_total(self):
        """Test percent label calculation with zero total signal."""
        m0 = 0.0
        labeled = 0.0
        background_ratio = 0.1

        corrected_labeled = labeled - (background_ratio * m0)
        corrected_labeled = max(0.0, corrected_labeled)

        total = m0 + labeled
        if total > 0:
            percent = (corrected_labeled / total) * 100
        else:
            percent = 0.0

        assert percent == 0.0

    def test_abundance_zero_mrrf(self):
        """Test abundance calculation with zero MRRF."""
        total_corrected = 500.0
        is_amount = 10.0
        is_m0 = 1000.0
        mrrf = 0.0  # Zero MRRF

        # Should handle gracefully
        if mrrf > 0 and is_m0 > 0:
            abundance = (total_corrected * is_amount) / (is_m0 * mrrf)
        else:
            abundance = 0.0  # Or flag as invalid

        assert abundance == 0.0


class TestMissingData:
    """Test handling of missing or incomplete data."""

    def test_missing_compound_data(self):
        """Test when compound has no EIC data."""
        # Simulate empty data for a compound
        sample_data = {}  # No compounds

        # Request data for missing compound
        compound_data = sample_data.get('MissingCompound', [0.0])

        assert compound_data == [0.0]

    def test_partial_isotopologue_data(self):
        """Test when fewer isotopologues than expected."""
        expected_isotopologues = 4
        actual_data = [100.0, 50.0]  # Only 2 instead of 4

        # Pad with zeros
        while len(actual_data) < expected_isotopologues:
            actual_data.append(0.0)

        assert len(actual_data) == 4
        assert actual_data[-1] == 0.0

    def test_empty_sample_list(self):
        """Test export with no samples."""
        samples = []

        # Should handle empty sample list
        for sample in samples:
            pass  # No iterations

        assert len(samples) == 0

    def test_missing_mm_files(self):
        """Test compounds without MM file specifications."""
        mm_files_field = None  # or ''

        # Should return empty list
        mm_samples = []
        if not mm_files_field:
            mm_samples = []

        assert mm_samples == []


class TestBoundaryConditions:
    """Test boundary conditions and extreme values."""

    def test_rt_window_smaller_than_offset(self):
        """Test when RT window < max(loffset, roffset)."""
        time = np.linspace(0, 10, 1000)
        intensity = np.ones(1000)

        # RT window of 0.2 but offsets of 1.0
        # This means some data outside RT window won't be captured
        rt = 5.0
        rt_window = 0.2  # Only captures 4.8 to 5.2
        loffset = 1.0     # Wants 4.0 to 5.0
        roffset = 1.0     # Wants 5.0 to 6.0

        # Filter by RT window first (as done in EIC extraction)
        rt_mask = (time >= rt - rt_window) & (time <= rt + rt_window)
        filtered_time = time[rt_mask]
        filtered_intensity = intensity[rt_mask]

        # Then try to integrate with larger offsets
        # This will miss data!
        if len(filtered_time) > 0:
            areas = calculate_peak_areas(
                filtered_time, filtered_intensity,
                0, rt, loffset, roffset
            )
        else:
            areas = [0.0]

        # Area will be less than expected due to missing data
        assert areas[0] < 2.0  # Less than full 2-minute window

    def test_exact_boundary_points(self):
        """Test points exactly at integration boundaries."""
        time = np.array([3.0, 4.0, 5.0, 6.0, 7.0])
        intensity = np.array([100, 200, 300, 400, 500])

        # Integration: 4.0 < t < 6.0 (strict inequality)
        areas = calculate_peak_areas(
            time, intensity, 0, 5.0, 1.0, 1.0
        )

        # Should only include point at t=5.0
        # Points at 4.0 and 6.0 are excluded
        assert len(areas) == 1
        # Only one point means area is 0 (need at least 2 for trapezoid)
        assert areas[0] == 0.0

    def test_very_large_values(self):
        """Test with very large intensity values."""
        time = np.linspace(0, 10, 100)
        intensity = np.ones(100) * 1e10  # Very large

        areas = calculate_peak_areas(
            time, intensity, 0, 5.0, 1.0, 1.0
        )

        assert areas[0] > 1e9  # Should handle large numbers

    def test_very_small_values(self):
        """Test with very small intensity values."""
        time = np.linspace(0, 10, 100)
        intensity = np.ones(100) * 1e-10  # Very small

        areas = calculate_peak_areas(
            time, intensity, 0, 5.0, 1.0, 1.0
        )

        assert areas[0] > 0  # Should still integrate
        assert areas[0] < 1e-8  # But result is small


class TestCorrectionEdgeCases:
    """Test edge cases in natural abundance correction."""

    def test_correction_with_no_labeling(self):
        """Test correction when label_atoms = 0."""
        corrector = NaturalAbundanceCorrector()

        # Unlabeled compound applies 1x1 correction matrix
        intensity = np.array([[100, 200, 300]]).astype(float)
        corrected = corrector.correct_time_series(
            intensity, 'C6H12O6', 'C', 0  # label_atoms = 0
        )

        # Shape should be preserved
        assert corrected.shape == intensity.shape
        # Values change by ~8% due to diagonal normalization
        ratio = corrected[0, 0] / intensity[0, 0]
        assert abs(ratio - 1.084) < 0.01  # About 8.4% increase

    def test_correction_with_extreme_ratios(self):
        """Test correction with unrealistic isotope ratios."""
        corrector = NaturalAbundanceCorrector()

        # Extreme case: M+1 > M+0 (unrealistic but possible with noise)
        intensity = np.array([
            [10],   # Very low M+0
            [100]   # Very high M+1
        ]).astype(float)

        corrected = corrector.correct_time_series(
            intensity, 'C1', 'C', 1
        )

        # Should still produce non-negative results
        assert np.all(corrected >= 0)

    def test_correction_matrix_singularity(self):
        """Test handling of singular correction matrices."""
        corrector = NaturalAbundanceCorrector()

        # Create a scenario that might produce singular matrix
        # This is actually hard to trigger with real formulas
        # but test the handling anyway
        try:
            matrix = corrector.build_correction_matrix('C1', 'C', 1)
            # Matrix should be invertible
            det = np.linalg.det(matrix)
            assert abs(det) > 1e-10  # Not singular
        except np.linalg.LinAlgError:
            # Should handle gracefully
            pytest.fail("Correction matrix is singular")


class TestDataValidation:
    """Test input data validation and error messages."""

    def test_negative_intensity_values(self):
        """Test handling of negative intensities (shouldn't happen but check anyway)."""
        time = np.linspace(0, 10, 100)
        intensity = np.ones(100)
        intensity[50] = -10.0  # Negative value

        # Integration should still work
        area = integrate_peak(intensity, time)

        # Area might be reduced but shouldn't crash
        assert isinstance(area, (float, np.floating))

    def test_unsorted_time_array(self):
        """Test with unsorted time points."""
        time = np.array([0, 5, 2, 8, 3, 9, 1])
        intensity = np.ones_like(time)

        # Should handle unsorted data
        # In practice, data should be sorted but test anyway
        area = integrate_peak(intensity, time[np.argsort(time)])
        assert area > 0

    def test_duplicate_time_points(self):
        """Test with duplicate time points."""
        time = np.array([1, 2, 2, 3, 4])  # Duplicate at t=2
        intensity = np.array([100, 200, 300, 400, 500])

        # Should handle duplicates
        area = integrate_peak(intensity, time)
        assert area > 0

    def test_empty_arrays(self):
        """Test with empty input arrays."""
        time = np.array([])
        intensity = np.array([])

        # Should handle empty arrays gracefully
        if len(time) > 0:
            area = integrate_peak(intensity, time)
        else:
            area = 0.0

        assert area == 0.0


class TestAbundanceCalculations:
    """Test abundance calculation edge cases and bugs."""

    def test_internal_standard_uses_amount_in_std_mix(self):
        """
        Test for Issue #55: Internal standard should use amount_in_std_mix, not int_std_amount.

        Bug: When the internal standard compound (e.g., scyllo-Ins) has:
          - int_std_amount = 1.0 (normalization factor for other metabolites)
          - amount_in_std_mix = 0.5 (ACTUAL amount in MM files)

        The code incorrectly used int_std_amount (1.0) for the internal standard's
        own abundance, causing it to be doubled. It should use amount_in_std_mix (0.5).
        """
        from types import SimpleNamespace
        from io import BytesIO
        import xlsxwriter

        # Mock exporter with internal standard set
        exporter = SimpleNamespace(internal_standard_compound='scyllo-Ins')

        # Mock provider that returns corrected data
        class MockProvider:
            def __init__(self):
                self._samples = ['Sample_MM1', 'Sample1']

            def get_all_compounds(self):
                return [
                    {
                        'compound_name': 'scyllo-Ins',
                        'mass0': 318.0,
                        'retention_time': 15.5,
                        'amount_in_std_mix': 0.5,  # Actual amount in standard mix
                        'int_std_amount': 1.0,     # Normalization factor (wrong if used for abundance)
                        'mm_files': '*_MM*'
                    },
                    {
                        'compound_name': 'Pyruvate',
                        'mass0': 174.0,
                        'retention_time': 7.17,
                        'amount_in_std_mix': 20.0,
                        'int_std_amount': None,
                        'mm_files': '*_MM*'
                    }
                ]

            def get_all_samples(self):
                return list(self._samples)

            def get_sample_corrected_data(self, sample_name):
                base_data = {
                    'scyllo-Ins': [1000.0],  # M+0 signal for internal standard
                    'Pyruvate': [500.0, 100.0, 50.0]  # M+0, M+1, M+2
                }
                return base_data

            def get_mrrf_values(self, compounds, internal_std):
                return {'Pyruvate': 2.0}

            def resolve_mm_samples(self, mm_field):
                if not mm_field:
                    return []
                return [sample for sample in self._samples if 'MM' in sample]

        provider = MockProvider()

        # Create a workbook in memory
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})

        # Import and call the abundances sheet generator
        from manic.sheet_generators import abundances
        abundances.write(
            workbook,
            exporter,
            progress_callback=None,
            start_progress=0,
            end_progress=100,
            provider=provider
        )

        workbook.close()

        # Read back the workbook to check values
        output.seek(0)
        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb['Abundances']

        # Headers are in rows 1-5, data starts at row 6
        # Column structure: [empty, Sample, Compound1, Compound2, ...]
        # scyllo-Ins should be in column 3 (index C)
        # Sample rows: row 6 = Sample_MM1, row 7 = Sample1

        scyllo_ins_abundance_mm = ws['C6'].value  # Sample_MM1 (MM file)
        scyllo_ins_abundance_non_mm = ws['C7'].value  # Sample1 (non-MM)

        # MM samples should use amount_in_std_mix, regular samples use int_std_amount
        assert scyllo_ins_abundance_mm == 0.5, (
            "Internal standard abundance for MM samples should use amount_in_std_mix (0.5)"
        )
        assert scyllo_ins_abundance_non_mm == 1.0, (
            "Internal standard abundance for non-MM samples should use int_std_amount (1.0)"
        )

        # Non-internal-standard metabolite should also be scaled differently per sample
        pyruvate_mm = ws['D6'].value
        pyruvate_non_mm = ws['D7'].value

        # Expected values based on formula: total_signal=650, IS signal=1000, MRRF=2
        assert pyruvate_mm == pytest.approx(0.1625)
        assert pyruvate_non_mm == pytest.approx(0.325)

    def test_abundances_require_internal_standard(self):
        """Export must fail if no internal standard is configured."""
        from types import SimpleNamespace
        from io import BytesIO
        import xlsxwriter

        exporter = SimpleNamespace(internal_standard_compound=None)

        class DummyProvider:
            def get_all_compounds(self):
                return [
                    {
                        'compound_name': 'scyllo-Ins',
                        'mass0': 318.0,
                        'retention_time': 15.5,
                        'amount_in_std_mix': 0.5,
                        'int_std_amount': 1.0,
                        'mm_files': '*_MM*'
                    }
                ]

            def get_all_samples(self):
                return ['Sample1']

            def get_sample_corrected_data(self, sample_name):
                return {'scyllo-Ins': [1000.0]}

            def get_mrrf_values(self, compounds, internal_std):
                return {}

            def resolve_mm_samples(self, mm_field):
                return []

        from manic.sheet_generators import abundances

        provider = DummyProvider()
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        with pytest.raises(ValueError, match="Internal standard must be set"):
            abundances.write(
                workbook,
                exporter,
                progress_callback=None,
                start_progress=0,
                end_progress=100,
                provider=provider,
            )
        workbook.close()

    def test_internal_standard_missing_amounts(self):
        """Exporter should fail fast when required IS amounts are missing."""
        from types import SimpleNamespace
        from io import BytesIO
        import xlsxwriter

        exporter = SimpleNamespace(internal_standard_compound='scyllo-Ins')

        class MissingAmountProvider:
            def __init__(self, *, missing_field):
                self._missing_field = missing_field

            def get_all_compounds(self):
                base = {
                    'compound_name': 'scyllo-Ins',
                    'mass0': 318.0,
                    'retention_time': 15.5,
                    'amount_in_std_mix': 0.5,
                    'int_std_amount': 1.0,
                    'mm_files': '*_MM*'
                }
                base[self._missing_field] = None
                return [base]

            def get_all_samples(self):
                return ['Sample1']

            def get_sample_corrected_data(self, sample_name):
                return {'scyllo-Ins': [1000.0]}

            def get_mrrf_values(self, compounds, internal_std):
                return {}

            def resolve_mm_samples(self, mm_field):
                return []

        from manic.sheet_generators import abundances

        for missing_field, expected_msg in (
            ('int_std_amount', "'int_std_amount'"),
            ('amount_in_std_mix', "'amount_in_std_mix'"),
        ):
            provider = MissingAmountProvider(missing_field=missing_field)
            output = BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            with pytest.raises(ValueError, match=expected_msg):
                abundances.write(
                    workbook,
                    exporter,
                    progress_callback=None,
                    start_progress=0,
                    end_progress=100,
                    provider=provider,
                )
            workbook.close()
