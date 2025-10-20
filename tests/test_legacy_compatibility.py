"""
Legacy compatibility and format tests for MANIC.
Tests for backward compatibility with MATLAB version, changelog formatting,
and legacy data import/export.
"""

from pathlib import Path

from openpyxl import Workbook

from manic.io.legacy_raw_values_reader import read_raw_values_workbook
from manic.io.in_memory_provider import InMemoryDataProvider
from manic.io.changelog_sections import (
    format_compounds_table_for_data_export,
    format_overrides_section_for_data_export,
    format_compounds_table_for_session_export,
    format_overrides_section_for_session_export,
)


# ============================================================================
# CHANGELOG FORMATTING TESTS
# ============================================================================

def test_format_compounds_table_for_data_export_single():
    """Test formatting compound table for data export changelog."""
    compounds = [
        {
            'compound_name': 'Glucose',
            'retention_time': 5.12345,
            'loffset': 0.5,
            'roffset': 0.75,
            'mass0': 180.063388,
            'label_atoms': 6,
            'formula': 'C6H12O6',
            'int_std_amount': 2.0,
        }
    ]
    out = format_compounds_table_for_data_export(compounds)

    # Headline and header row
    assert '## Compounds Processed' in out
    assert '| Compound Name | RT (min) | L Offset | R Offset | Mass (m/z) | Label Atoms | Formula | Internal Std Amount |' in out
    # Row content with precision
    assert '| Glucose | 5.123 | 0.500 | 0.750 | 180.0634 | 6 | C6H12O6 | 2.0 |' in out


def test_format_overrides_section_for_data_export_multiple():
    """Test formatting overrides section with multiple entries."""
    overrides = [
        {'compound_name': 'Glucose', 'sample_name': 'S1', 'retention_time': 5.0, 'loffset': 0.5, 'roffset': 0.7},
        {'compound_name': 'Lactate', 'sample_name': 'S2', 'retention_time': 3.2, 'loffset': 0.3, 'roffset': 0.4},
    ]
    out = format_overrides_section_for_data_export(overrides)
    assert '## Session Parameter Overrides' in out
    assert '| Glucose | S1 | 5.000 | 0.500 | 0.700 |' in out
    assert '| Lactate | S2 | 3.200 | 0.300 | 0.400 |' in out


def test_format_compounds_table_for_session_export_single():
    """Test formatting compound table for session export."""
    compounds = [
        {
            'compound_name': 'Alanine',
            'retention_time': 2.4567,
            'loffset': 0.2,
            'roffset': 0.3,
            'mass0': 89.0477,
            'label_atoms': 3,
        }
    ]
    out = format_compounds_table_for_session_export(compounds)
    assert '## Compound Definitions (1 compounds)' in out
    assert '| Alanine | 2.457 | 0.200 | 0.300 | 89.0477 | 3 |' in out


def test_format_overrides_section_for_session_export_grouping():
    """Test formatting overrides section with compound grouping."""
    overrides = [
        {'compound_name': 'Alanine', 'sample_name': 'S1', 'retention_time': 2.1, 'loffset': 0.2, 'roffset': 0.3},
        {'compound_name': 'Alanine', 'sample_name': 'S0', 'retention_time': 2.2, 'loffset': 0.25, 'roffset': 0.35},
        {'compound_name': 'Citrate', 'sample_name': 'S2', 'retention_time': 7.3, 'loffset': 0.6, 'roffset': 0.8},
    ]
    out = format_overrides_section_for_session_export(overrides)
    # Section header with count
    assert '## Session Integration Overrides (3 overrides)' in out
    # Group headings present
    assert '### Alanine' in out
    assert '### Citrate' in out
    # Rows with precision
    assert '| S0 | 2.200 | 0.250 | 0.350 |' in out
    assert '| S1 | 2.100 | 0.200 | 0.300 |' in out
    assert '| S2 | 7.300 | 0.600 | 0.800 |' in out


# ============================================================================
# LEGACY RAW VALUES IMPORT/EXPORT TESTS
# ============================================================================

def write_raw_values_xlsx(path: str):
    """Helper to create a raw values Excel file in legacy format."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Raw Values'

    # Headers
    ws.cell(row=1, column=1, value='Compound Name')
    ws.cell(row=2, column=1, value='Mass')
    ws.cell(row=3, column=1, value='Isotope')
    ws.cell(row=4, column=1, value='tR')
    ws.cell(row=1, column=2, value=None)
    ws.cell(row=2, column=2, value=None)
    ws.cell(row=3, column=2, value=None)
    ws.cell(row=4, column=2, value=None)

    # Two compounds: A with 1 label (M+0,M+1), B with unlabeled only
    ws.cell(row=1, column=3, value='A')
    ws.cell(row=2, column=3, value=100.0)
    ws.cell(row=3, column=3, value=0)
    ws.cell(row=4, column=3, value=1.0)

    ws.cell(row=1, column=4, value='A')
    ws.cell(row=2, column=4, value=100.0)
    ws.cell(row=3, column=4, value=1)
    ws.cell(row=4, column=4, value=1.0)

    ws.cell(row=1, column=5, value='B')
    ws.cell(row=2, column=5, value=50.0)
    ws.cell(row=3, column=5, value=0)
    ws.cell(row=4, column=5, value=2.0)

    # Sample rows
    ws.cell(row=5, column=1, value=None)
    ws.cell(row=5, column=2, value='S1')
    ws.cell(row=5, column=3, value=10.0)  # A M+0
    ws.cell(row=5, column=4, value=2.0)   # A M+1
    ws.cell(row=5, column=5, value=5.0)   # B M+0

    ws.cell(row=6, column=1, value=None)
    ws.cell(row=6, column=2, value='S2')
    ws.cell(row=6, column=3, value=20.0)
    ws.cell(row=6, column=4, value=3.0)
    ws.cell(row=6, column=5, value=7.0)

    wb.save(path)


def test_read_raw_values_workbook_roundtrip(tmp_path: Path):
    """Test reading raw values from legacy Excel format."""
    xlsx_path = tmp_path / 'raw_values.xlsx'
    write_raw_values_xlsx(str(xlsx_path))
    samples, raw = read_raw_values_workbook(str(xlsx_path))
    assert samples == ['S1', 'S2']
    assert raw['S1']['A'] == [10.0, 2.0]
    assert raw['S1']['B'] == [5.0]
    assert raw['S2']['A'] == [20.0, 3.0]
    assert raw['S2']['B'] == [7.0]


# ============================================================================
# IN-MEMORY PROVIDER TESTS (LEGACY REBUILD)
# ============================================================================

def test_in_memory_provider_corrected_vector(tmp_path: Path):
    """Test in-memory provider for legacy rebuild functionality."""
    # Compounds: A has label_atoms=1 (2 iso), B is unlabeled
    compounds = [
        {
            'compound_name': 'A',
            'retention_time': 1.0,
            'mass0': 100.0,
            'loffset': 0.1,
            'roffset': 0.1,
            'label_atoms': 1,
            'formula': 'C1H4',
            'label_type': 'C',
            'tbdms': 0,
            'meox': 0,
            'me': 0,
        },
        {
            'compound_name': 'B',
            'retention_time': 2.0,
            'mass0': 50.0,
            'loffset': 0.1,
            'roffset': 0.1,
            'label_atoms': 0,
            'formula': None,
            'label_type': 'C',
            'tbdms': 0,
            'meox': 0,
            'me': 0,
        },
    ]
    samples = ['S1']
    raw = {'S1': {'A': [10.0, 2.0], 'B': [5.0]}}

    prov = InMemoryDataProvider(compounds, samples, raw)

    corr = prov.get_sample_corrected_data('S1')
    # A: two isotopologues, non-negative, some total (approximate mode)
    assert isinstance(corr['A'], list) and len(corr['A']) == 2
    assert all(v >= 0 for v in corr['A'])
    # B: unlabeled, copied as-is
    assert corr['B'] == [5.0]