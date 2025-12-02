import pytest
from types import SimpleNamespace
from io import BytesIO
import xlsxwriter
from openpyxl import load_workbook

from manic.sheet_generators import carbon_enrichment


class TestCarbonEnrichmentAPE:
    """Tests for Atom Percent Excess (APE) calculation with background subtraction."""

    @pytest.fixture
    def mock_provider_with_background(self):
        """
        Provider with MM baseline for background subtraction testing.

        Scenario: Background (MM) is 50% enriched for Glucose.
        - MM_01: 50% M+0, 50% M+6 -> 50% enrichment
        - Sample_Baseline: Same as MM -> Expect 0% APE
        - Sample_High: 100% M+6 -> 100% enrichment -> Expect 50% APE
        - Sample_Low: 100% M+0 -> 0% enrichment -> Expect 0% APE (clamped from -50%)
        """

        class MockProvider:
            def get_all_compounds(self):
                return [
                    {
                        "compound_name": "Glucose",
                        "label_atoms": 6,
                        "mass0": 100.0,
                        "retention_time": 5.0,
                        "mm_files": "*MM*",
                    }
                ]

            def get_all_samples(self):
                return ["MM_01", "Sample_Baseline", "Sample_High", "Sample_Low"]

            def resolve_mm_samples(self, pattern):
                return ["MM_01"]

            def get_sample_corrected_data(self, sample_name):
                # Glucose (6C): Background is 50% enriched
                # [M+0, M+1, M+2, M+3, M+4, M+5, M+6]
                # Weighted Sum = (0*50 + 6*50) = 300. Max = 6*100 = 600. Enrichment = 50%.
                background_data = [50.0, 0.0, 0.0, 0.0, 0.0, 0.0, 50.0]

                if sample_name == "MM_01":
                    return {"Glucose": background_data}

                elif sample_name == "Sample_Baseline":
                    # Exact same as background -> Expect 0% Excess
                    return {"Glucose": background_data}

                elif sample_name == "Sample_High":
                    # 100% Fully Labelled -> 100% Enrichment
                    # Excess = 100% - 50% (background) = 50%
                    return {"Glucose": [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 100.0]}

                elif sample_name == "Sample_Low":
                    # 100% Unlabelled -> 0% Enrichment
                    # Excess = 0% - 50% (background) = -50% -> Clamp to 0.0
                    return {"Glucose": [100.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]}

                return {}

        return MockProvider()

    def test_background_subtraction(self, mock_provider_with_background):
        """Verify APE calculation correctly subtracts MM baseline."""
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        exporter = SimpleNamespace()

        carbon_enrichment.write(
            workbook, exporter, None, 0, 100, provider=mock_provider_with_background
        )
        workbook.close()

        output.seek(0)
        wb = load_workbook(output)
        ws = wb["% Carbons Labelled"]

        # MM_01 (Row 5): Should be 0.0 (Background - Background)
        assert ws["C5"].value == 0.0

        # Sample_Baseline (Row 6): Should be 0.0 (50% - 50%)
        assert ws["C6"].value == 0.0

        # Sample_High (Row 7): Should be 50.0 (100% - 50%)
        assert ws["C7"].value == 50.0

        # Sample_Low (Row 8): Should be 0.0 (0% - 50% clamped)
        assert ws["C8"].value == 0.0

    @pytest.fixture
    def mock_provider_no_background(self):
        """Provider without MM files - should report absolute enrichment."""

        class MockProvider:
            def get_all_compounds(self):
                return [
                    {
                        "compound_name": "Alanine",
                        "label_atoms": 3,
                        "mass0": 50.0,
                        "retention_time": 2.5,
                        "mm_files": None,  # No MM pattern
                    }
                ]

            def get_all_samples(self):
                return ["Sample_Full"]

            def resolve_mm_samples(self, pattern):
                return []  # No MM samples found

            def get_sample_corrected_data(self, sample_name):
                # 100% M+3 for Alanine (3C) -> 100% enrichment
                return {"Alanine": [0.0, 0.0, 0.0, 1000.0]}

        return MockProvider()

    def test_no_mm_files_uses_zero_baseline(self, mock_provider_no_background):
        """When no MM files exist, baseline is 0 (absolute enrichment reported)."""
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        exporter = SimpleNamespace()

        carbon_enrichment.write(
            workbook, exporter, None, 0, 100, provider=mock_provider_no_background
        )
        workbook.close()

        output.seek(0)
        wb = load_workbook(output)
        ws = wb["% Carbons Labelled"]

        # Sample_Full: 100% enrichment, 0 baseline = 100% APE
        assert ws["C5"].value == 100.0


class TestCarbonEnrichmentCalculation:
    """Tests for the enrichment calculation helper function."""

    def test_fully_unlabelled(self):
        """100% M+0 should give 0% enrichment."""
        data = [1000.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        result = carbon_enrichment.calculate_enrichment(data, label_atoms=6)
        assert result == 0.0

    def test_fully_labelled(self):
        """100% M+N should give 100% enrichment."""
        # 6-carbon compound, all M+6
        data = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 1000.0]
        result = carbon_enrichment.calculate_enrichment(data, label_atoms=6)
        assert result == 100.0

    def test_single_label(self):
        """100% M+1 for 6C compound should give 16.67% enrichment."""
        data = [0.0, 1000.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        result = carbon_enrichment.calculate_enrichment(data, label_atoms=6)
        assert result == pytest.approx(16.666666, rel=1e-4)

    def test_mixed_distribution(self):
        """50% M+0, 50% M+6 should give 50% enrichment."""
        data = [500.0, 0.0, 0.0, 0.0, 0.0, 0.0, 500.0]
        result = carbon_enrichment.calculate_enrichment(data, label_atoms=6)
        assert result == 50.0

    def test_zero_label_atoms(self):
        """Zero label atoms should return 0 (avoid division by zero)."""
        data = [1000.0, 500.0]
        result = carbon_enrichment.calculate_enrichment(data, label_atoms=0)
        assert result == 0.0

    def test_empty_data(self):
        """Empty/zero data should return 0."""
        result = carbon_enrichment.calculate_enrichment([0.0], label_atoms=6)
        assert result == 0.0

        result = carbon_enrichment.calculate_enrichment([], label_atoms=6)
        assert result == 0.0


class TestCarbonEnrichmentEdgeCases:
    """Tests for edge cases and validation."""

    @pytest.fixture
    def mock_provider_with_unlabelable(self):
        """Provider with unlabelable compound (label_atoms=0)."""

        class MockProvider:
            def get_all_compounds(self):
                return [
                    {
                        "compound_name": "Unlabelable",
                        "label_atoms": 0,
                        "mass0": 200.0,
                        "retention_time": 10.0,
                        "mm_files": "*MM*",
                    }
                ]

            def get_all_samples(self):
                return ["MM_01", "Sample_A"]

            def resolve_mm_samples(self, pattern):
                return ["MM_01"]

            def get_sample_corrected_data(self, sample_name):
                return {"Unlabelable": [1000.0]}

        return MockProvider()

    def test_unlabelable_compound(self, mock_provider_with_unlabelable):
        """Compounds with label_atoms=0 should always report 0%."""
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        exporter = SimpleNamespace()

        carbon_enrichment.write(
            workbook, exporter, None, 0, 100, provider=mock_provider_with_unlabelable
        )
        workbook.close()

        output.seek(0)
        wb = load_workbook(output)
        ws = wb["% Carbons Labelled"]

        # Both samples should be 0.0 for unlabelable compound
        assert ws["C5"].value == 0.0
        assert ws["C6"].value == 0.0

    @pytest.fixture
    def mock_provider_multiple_mm(self):
        """Provider with multiple MM samples for averaging."""

        class MockProvider:
            def get_all_compounds(self):
                return [
                    {
                        "compound_name": "Glucose",
                        "label_atoms": 6,
                        "mass0": 100.0,
                        "retention_time": 5.0,
                        "mm_files": "*MM*",
                    }
                ]

            def get_all_samples(self):
                return ["MM_01", "MM_02", "Sample_Test"]

            def resolve_mm_samples(self, pattern):
                return ["MM_01", "MM_02"]

            def get_sample_corrected_data(self, sample_name):
                if sample_name == "MM_01":
                    # 100% M+0 -> 0% enrichment
                    return {"Glucose": [100.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]}
                elif sample_name == "MM_02":
                    # 100% M+6 -> 100% enrichment
                    return {"Glucose": [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 100.0]}
                elif sample_name == "Sample_Test":
                    # 100% M+6 -> 100% enrichment
                    # Baseline avg = (0 + 100) / 2 = 50%
                    # APE = 100 - 50 = 50%
                    return {"Glucose": [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 100.0]}
                return {}

        return MockProvider()

    def test_multiple_mm_averaging(self, mock_provider_multiple_mm):
        """Multiple MM samples should be averaged for baseline."""
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        exporter = SimpleNamespace()

        carbon_enrichment.write(
            workbook, exporter, None, 0, 100, provider=mock_provider_multiple_mm
        )
        workbook.close()

        output.seek(0)
        wb = load_workbook(output)
        ws = wb["% Carbons Labelled"]

        # Sample_Test (Row 7): 100% enrichment - 50% avg baseline = 50%
        assert ws["C7"].value == 50.0
